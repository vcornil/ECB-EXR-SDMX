import sdmx
import pandas as pd
from datetime import datetime, date, timedelta
import numpy as np
import pprint
import argparse
from tabulate import tabulate
from fpdf import FPDF
from fpdf.enums import XPos, YPos, Align
import csv
from openpyxl import Workbook
from typing import Dict, Union, cast
import os
from pathlib import Path
import sys

class Ecbexr:
    # Class constants
    # Default to IMF provider; can be overridden per-instance
    SDMX_CLIENT = "IMF_DATA"
    SDMX_DATASET = "IFS"
    SDMX_DAILY_KEY = None  # IMF typically exposes monthly series; set to None to skip daily by default
    SDMX_MONTHLY_KEY = None

    def __init__(
        self,
        target_year: int | None = None,
        target_month: int | None = None,
        from_month: int | None = None,
        provider: str | None = None,
        dataset: str | None = None,
        daily_key: str | None = None,
        monthly_key: str | None = None,
    ):

        # initialize internal variables
        self._start_month = 1
        self._end_month = 1
        self._end_year = 2025
        self._start_year = 2025

        # Default values for target year is current year iif month is not January else previous year if no value provided
        if target_year == None:
            self.end_year = (
                datetime.now().year
                if datetime.now().month > 1
                else datetime.now().year - 1
            )
        else:
            self.end_year = target_year

        # Default values for target month is previous month iif month is not January else December if no value provided
        if target_month == None:
            self.end_month = (
                datetime.now().month - 1 if datetime.now().month > 1 else 12
            )
        else:
            self.end_month = target_month

        # Default values for starting month is January if no value provided
        if from_month == None:
            self.start_month = 1
        else:
            self.start_month = from_month

        # Default values for starting year is _target_year if start month if lower than end month
        self.start_year = (
            self._end_year
            if self.start_month <= self.end_month
            else (self._end_year - 1)
        )

        # SDMX provider / dataset / key overrides (useful to target IMF, ECB, etc.)
        self.sdmx_provider = provider or Ecbexr.SDMX_CLIENT
        self.sdmx_dataset = dataset or Ecbexr.SDMX_DATASET
        # If daily_key is None, daily retrieval will be skipped
        self.sdmx_daily_key = (
            daily_key if daily_key is not None else Ecbexr.SDMX_DAILY_KEY
        )
        self.sdmx_monthly_key = (
            monthly_key if monthly_key is not None else Ecbexr.SDMX_MONTHLY_KEY
        )

    # Getter for the end year
    @property
    def end_year(self) -> int:
        return self._end_year

    # Setter for the end year
    @end_year.setter
    def end_year(self, value: int):
        """
        Setter for the year property with type and range validation.
        """
        if not isinstance(value, int):
            raise TypeError("End Year must be an integer.")
        if not 2001 <= value <= datetime.now().year:
            raise ValueError(
                f"End Year must be between 2001 and {datetime.now().year}."
            )
        self._end_year = value

    # Getter for the end month
    @property
    def end_month(self) -> int:
        return self._end_month

    # Setter for the end month
    @end_month.setter
    def end_month(self, value: int) -> None:
        """
        Setter for the end month property with type and range validation.
        """
        if not isinstance(value, int):
            raise TypeError("End Month must be an integer.")
        if not 1 <= value <= 12:
            raise ValueError("End Month must be between 1 and 12.")
        self._end_month = value
        # Also reset start year, in case it is needed
        self.start_year = (
            self._end_year if self.start_month <= value else (self._end_year - 1)
        )

    # Getter for the start month
    @property
    def start_month(self) -> int:
        return self._start_month

    # Setter for the start month
    @start_month.setter
    def start_month(self, value: int) -> None:
        """
        Setter for the start month property with type and range validation.
        """
        if not isinstance(value, int):
            raise TypeError("Start Month must be an integer.")
        if not 1 <= value <= 12:
            raise ValueError("Start Month must be between 1 and 12.")
        self._start_month = value
        # Also reset start year, in case it is needed
        self.start_year = (
            self._end_year if value <= self.end_month else (self._end_year - 1)
        )

    def rates(self) -> dict[str, dict[str, float | str | None]]:
        """
        Retrieve exchange rates for a specific month and compute:
        1. Month-end closing rate (from daily data - last trading day of the month)
        2. YTD average rate (average of monthly rates from Jan to target month)
        3. Target month average (official ECB monthly average for the specific month)

        Returns:
            dict: Dictionary with currency codes as keys and rate info as values
        """

        # Set the Client (provider can be e.g. 'ECB', 'IMF_DATA' or a base URL known by sdmx)
        client = sdmx.Client(self.sdmx_provider)

        try:
            # Step 1: Get daily data ONLY for the target month to get month-end closing rate - limiting the number of days as only the last quotation day is needed
            if self.end_month == 12:
                last_day = 31
            else:
                next_month = datetime(self.end_year, self.end_month + 1, 1)
                last_day_date = next_month - timedelta(days=1)
                last_day = last_day_date.day

            daily_start = f"{self.end_year}-{self.end_month:02d}-{(last_day-6):02d}"  # at least 1 week of data to tke weekends and holidays into account
            daily_end = f"{self.end_year}-{self.end_month:02d}-{last_day:02d}"

            # print(f"Fetching daily data for month-end rates: {daily_start} to {daily_end}...")
            if self.sdmx_daily_key:
                daily_response = client.data(
                    self.sdmx_dataset,
                    key=self.sdmx_daily_key,
                    params={"startPeriod": daily_start, "endPeriod": daily_end},
                )

                daily_df = sdmx.to_pandas(daily_response.data)
                # normalize Series -> DataFrame
                if isinstance(daily_df, pd.Series):
                    daily_df = daily_df.reset_index().rename(columns={0: 'value'})
            else:
                # No daily key provided: attempt to fetch the dataset without a series key
                # (this will retrieve all series for the dataset within the period)
                try:
                    daily_response = client.data(
                        self.sdmx_dataset,
                        params={"startPeriod": daily_start, "endPeriod": daily_end},
                    )
                    daily_df = sdmx.to_pandas(daily_response.data)
                    if isinstance(daily_df, pd.Series):
                        daily_df = daily_df.reset_index().rename(columns={0: 'value'})
                except Exception:
                    # Provider may not support daily data or the request may be too large
                    daily_df = pd.DataFrame()

            # Step 2: Get monthly data from start month+year to target month for averages
            monthly_start = f"{self.start_year}-{self.start_month:02d}"
            monthly_end = f"{self.end_year}-{self.end_month:02d}"

            # print(f"Fetching monthly data for averages: {monthly_start} to {monthly_end}...")
            if self.sdmx_monthly_key:
                monthly_response = client.data(
                    self.sdmx_dataset,
                    key=self.sdmx_monthly_key,
                    params={"startPeriod": monthly_start, "endPeriod": monthly_end},
                )

                monthly_df = sdmx.to_pandas(monthly_response.data)
                if isinstance(monthly_df, pd.Series):
                    monthly_df = monthly_df.reset_index().rename(columns={0: 'value'})
            else:
                # No monthly key provided: fetch all series for the dataset (may be large)
                try:
                    monthly_response = client.data(
                        self.sdmx_dataset,
                        params={"startPeriod": monthly_start, "endPeriod": monthly_end},
                    )
                    monthly_df = sdmx.to_pandas(monthly_response.data)
                    if isinstance(monthly_df, pd.Series):
                        monthly_df = monthly_df.reset_index().rename(columns={0: 'value'})
                except Exception:
                    monthly_df = pd.DataFrame()

            # If both daily and monthly data are empty, nothing to do
            if daily_df.empty and monthly_df.empty:
                print("No valid data retrieved.")
                return {}

            # print(f"Daily data shape: {daily_df.shape}")
            # print(f"Monthly data shape: {monthly_df.shape}")
            # print(daily_df.items)

            # Process daily data for month-end rates
            closing_rates: dict[str, dict[str, float | str | None]] = {}
            if not daily_df.empty:
                daily_df = daily_df.reset_index()

                if "TIME_PERIOD" in daily_df.columns:
                    daily_df["date"] = pd.to_datetime(daily_df["TIME_PERIOD"])
                else:
                    date_cols = [
                        col
                        for col in daily_df.columns
                        if "TIME" in col.upper() or "DATE" in col.upper()
                    ]
                    if date_cols:
                        daily_df["date"] = pd.to_datetime(daily_df[date_cols[0]])

                # Identify currency and value columns
                daily_currency_col = None
                for col in daily_df.columns:
                    if "CURRENCY" in col.upper() and "DENOM" not in col.upper():
                        daily_currency_col = col
                        break

                daily_value_col = (
                    "value"
                    if "value" in daily_df.columns
                    else daily_df.select_dtypes(include=[np.number]).columns[-1]
                )

                if daily_currency_col:
                    # print(f"Daily - Currency column: {daily_currency_col}, Value column: {daily_value_col}")

                    # Get month-end rate for each currency (last trading day)
                    daily_df = daily_df.sort_values(["date", daily_currency_col])

                    for currency in daily_df[daily_currency_col].unique():
                        if pd.isna(currency):
                            continue

                        currency_daily = daily_df[
                            daily_df[daily_currency_col] == currency
                        ].copy()
                        currency_daily = currency_daily.dropna(subset=[daily_value_col])

                        if len(currency_daily) > 0:
                            # Get the last (most recent) rate in the month -> last row -> iloc[-1]
                            closing_rate: float | None = float(
                                currency_daily.iloc[-1][daily_value_col]
                            )
                            closing_date: str | None = currency_daily.iloc[-1][
                                "date"
                            ].strftime("%Y-%m-%d")

                            # store the closing rate
                            closing_rates[currency] = {
                                "rate": closing_rate,
                                "date": closing_date,
                            }

            # Process monthly data for averages
            monthly_averages = {}
            ytd_averages = {}

            if not monthly_df.empty:
                monthly_df = monthly_df.reset_index()

                if "TIME_PERIOD" in monthly_df.columns:
                    # IMF uses format like '2025-M01', '2025-M10'
                    # Parse more carefully for IMF format
                    def parse_imf_period(period_str):
                        try:
                            # Format: YYYY-M## e.g., 2025-M10
                            if isinstance(period_str, str) and 'M' in period_str:
                                year, month_str = period_str.split('-M')
                                month = int(month_str)
                                return pd.Timestamp(year=int(year), month=month, day=1)
                            else:
                                return pd.to_datetime(period_str)
                        except Exception:
                            return pd.to_datetime(period_str)
                    
                    monthly_df["month_period"] = monthly_df["TIME_PERIOD"].apply(parse_imf_period)
                else:
                    date_cols = [
                        col
                        for col in monthly_df.columns
                        if "TIME" in col.upper() or "DATE" in col.upper()
                    ]
                    if date_cols:
                        monthly_df["month_period"] = pd.to_datetime(
                            monthly_df[date_cols[0]]
                        )

                # Identify currency and value columns (different sources have different structures)
                monthly_currency_col = None
                indicator_col = None
                country_col = None
                for col in monthly_df.columns:
                    if "CURRENCY" in col.upper() and "DENOM" not in col.upper():
                        monthly_currency_col = col
                    elif col.upper() == "INDICATOR":
                        indicator_col = col
                    elif col.upper() == "COUNTRY":
                        country_col = col

                monthly_value_col = (
                    "value"
                    if "value" in monthly_df.columns
                    else monthly_df.select_dtypes(include=[np.number]).columns[-1]
                )

                # Filter to current year and up to target month (for all sources)
                monthly_df = monthly_df[
                    (
                        (monthly_df["month_period"].dt.year == self.end_year)
                        & (monthly_df["month_period"].dt.month <= self.end_month)
                    )
                    | (
                        (monthly_df["month_period"].dt.year == self.start_year)
                        & (monthly_df["month_period"].dt.month >= self.start_month)
                    )
                ]

                # ECB format: use CURRENCY column
                if monthly_currency_col:
                    # print(f"Monthly - Currency column: {monthly_currency_col}, Value column: {monthly_value_col}")
                    for currency in monthly_df[monthly_currency_col].unique():
                        if pd.isna(currency):
                            continue

                        currency_monthly = monthly_df[
                            monthly_df[monthly_currency_col] == currency
                        ].copy()
                        currency_monthly = currency_monthly.dropna(
                            subset=[monthly_value_col]
                        )
                        currency_monthly = currency_monthly.sort_values("month_period")

                        if len(currency_monthly) == 0:
                            continue

                        # Get target month average
                        target_month_data = currency_monthly[
                            currency_monthly["month_period"].dt.month == self.end_month
                        ]

                        if len(target_month_data) > 0:
                            monthly_averages[currency] = float(
                                target_month_data.iloc[0][monthly_value_col]
                            )

                        # Calculate YTD average (average of monthly averages)
                        if len(currency_monthly) > 0:
                            ytd_averages[currency] = float(
                                currency_monthly[monthly_value_col].mean()
                            )

                # IMF ER format: country-indexed exchange rates
                elif indicator_col and country_col:
                    # IMF ER dataset: each COUNTRY has exchange rates (EUR_XDC, USD_XDC, etc.)
                    # Map country codes to standard 3-letter currency codes
                    # Source: ISO 4217 currency codes for IMF member countries
                    country_to_currency = {
                        'ABW': 'AWG', 'AFG': 'AFN', 'AGO': 'AOA', 'AIA': 'XCD', 'ALA': 'EUR',
                        'ALB': 'ALL', 'AND': 'EUR', 'ARE': 'AED', 'ARG': 'ARS', 'ARM': 'AMD',
                        'ATG': 'XCD', 'AUS': 'AUD', 'AUT': 'EUR', 'AZE': 'AZN',
                        'BDI': 'BIF', 'BEL': 'EUR', 'BEN': 'XOF', 'BES': 'USD', 'BFA': 'XOF',
                        'BGD': 'BDT', 'BGR': 'BGN', 'BHR': 'BHD', 'BHS': 'BSD', 'BIH': 'BAM',
                        'BLR': 'BYN', 'BLZ': 'BZD', 'BMU': 'BMD', 'BOL': 'BOB', 'BRA': 'BRL',
                        'BRB': 'BBD', 'BRN': 'BND', 'BTN': 'BTN', 'BWA': 'BWP', 'CAF': 'XAF',
                        'CAN': 'CAD', 'CHE': 'CHF', 'CHL': 'CLP', 'CHN': 'CNY', 'CIV': 'XOF',
                        'CMR': 'XAF', 'COL': 'COP', 'COM': 'KMF', 'CPV': 'CVE', 'CRI': 'CRC',
                        'CUB': 'CUP', 'CUW': 'ANG', 'CYP': 'EUR', 'CZE': 'CZK',
                        'DEU': 'EUR', 'DJI': 'DJF', 'DMA': 'XCD', 'DNK': 'DKK', 'DOM': 'DOP',
                        'DZA': 'DZD',
                        'ECU': 'USD', 'EGY': 'EGP', 'ERI': 'ERN', 'ESH': 'MAD', 'ESP': 'EUR',
                        'EST': 'EUR', 'ETH': 'ETB',
                        'FIN': 'EUR', 'FJI': 'FJD', 'FLK': 'FKP', 'FRA': 'EUR', 'FRO': 'DKK',
                        'FSM': 'USD',
                        'GAB': 'XAF', 'GBR': 'GBP', 'GEO': 'GEL', 'GHA': 'GHS', 'GIB': 'GIP',
                        'GRD': 'XCD', 'GRC': 'EUR', 'GRL': 'DKK', 'GTM': 'GTQ', 'GUM': 'USD',
                        'GUY': 'GYD',
                        'HKG': 'HKD', 'HND': 'HNL', 'HRV': 'HRK', 'HTI': 'HTG', 'HUN': 'HUF',
                        'IDN': 'IDR', 'IND': 'INR', 'IRL': 'EUR', 'IRN': 'IRR', 'IRQ': 'IQD',
                        'ISL': 'ISK', 'ISR': 'ILS', 'ITA': 'EUR',
                        'JAM': 'JMD', 'JOR': 'JOD', 'JPN': 'JPY',
                        'KAZ': 'KZT', 'KEN': 'KES', 'KGZ': 'KGS', 'KHM': 'KHR', 'KIR': 'AUD',
                        'KNA': 'XCD', 'KOR': 'KRW', 'KWT': 'KWD',
                        'LAO': 'LAK', 'LBN': 'LBP', 'LBR': 'LRD', 'LBY': 'LYD', 'LCA': 'XCD',
                        'LIE': 'CHF', 'LKA': 'LKR', 'LSO': 'LSL', 'LTU': 'EUR', 'LUX': 'EUR',
                        'LVA': 'EUR',
                        'MAC': 'MOP', 'MAF': 'ANG', 'MAR': 'MAD', 'MCO': 'EUR', 'MDA': 'MDL',
                        'MDG': 'MGA', 'MEX': 'MXN', 'MHL': 'USD', 'MKD': 'MKD', 'MLI': 'XOF',
                        'MLT': 'EUR', 'MMR': 'MMK', 'MNE': 'EUR', 'MNG': 'MNT', 'MNP': 'USD',
                        'MOZ': 'MZN', 'MRT': 'MRU', 'MUS': 'MUR', 'MDV': 'MVR', 'MWI': 'MWK',
                        'MYS': 'MYR',
                        'NAM': 'NAD', 'NCL': 'XPF', 'NER': 'XOF', 'NFK': 'AUD', 'NGA': 'NGN',
                        'NIC': 'NIO', 'NLD': 'EUR', 'NOR': 'NOK', 'NPL': 'NPR', 'NRU': 'AUD',
                        'NZL': 'NZD',
                        'OMN': 'OMR',
                        'PAK': 'PKR', 'PAN': 'PAB', 'PCN': 'NZD', 'PER': 'PEN', 'PHL': 'PHP',
                        'PLW': 'USD', 'PNG': 'PGK', 'POL': 'PLN', 'PRI': 'USD', 'PRK': 'KPW',
                        'PRT': 'EUR', 'PRY': 'PYG',
                        'QAT': 'QAR',
                        'ROU': 'RON', 'RUS': 'RUB', 'RWA': 'RWF',
                        'SAU': 'SAR', 'SDN': 'SDG', 'SEN': 'XOF', 'SGP': 'SGD', 'SHN': 'SHP',
                        'SJM': 'NOK', 'SVK': 'EUR', 'SVN': 'EUR', 'SWE': 'SEK', 'SWZ': 'SZL',
                        'SXM': 'ANG', 'SYC': 'SCR', 'SYR': 'SYP',
                        'TCA': 'USD', 'TCD': 'XAF', 'TGO': 'XOF', 'THA': 'THB', 'TJK': 'TJS',
                        'TKL': 'NZD', 'TKM': 'TMT', 'TLS': 'USD', 'TON': 'TOP', 'TTO': 'TTD',
                        'TUN': 'TND', 'TUR': 'TRY', 'TUV': 'AUD', 'TWN': 'TWD', 'TZA': 'TZS',
                        'UGA': 'UGX', 'UKR': 'UAH', 'URY': 'UYU', 'USA': 'USD', 'UZB': 'UZS',
                        'VAT': 'EUR', 'VCT': 'XCD', 'VEN': 'VES', 'VGB': 'USD', 'VIR': 'USD',
                        'VNM': 'VND',
                        'VUT': 'VUV',
                        'WLF': 'XPF', 'WSM': 'WST',
                        'YEM': 'YER',
                        'ZAF': 'ZAR', 'ZMB': 'ZMW', 'ZWE': 'ZWL',
                    }
                    
                    for country in monthly_df[country_col].unique():
                        if pd.isna(country):
                            continue
                        
                        # Map country code to currency code
                        currency_code = country_to_currency.get(country, country)
                        
                        # Get all data for this country
                        country_data = monthly_df[monthly_df[country_col] == country].copy()
                        
                        # For IMF, we'll use USD_XDC rates (USD per country's currency)
                        # Filter to USD_XDC indicator (USD as base currency)
                        usd_xdc_data = country_data[country_data[indicator_col] == 'USD_XDC'].copy()
                        
                        if len(usd_xdc_data) == 0:
                            continue
                        
                        # Separate period average and end of period
                        if 'TYPE_OF_TRANSFORMATION' in usd_xdc_data.columns:
                            pa_data = usd_xdc_data[usd_xdc_data['TYPE_OF_TRANSFORMATION'] == 'PA_RT'].copy()
                            eop_data = usd_xdc_data[usd_xdc_data['TYPE_OF_TRANSFORMATION'] == 'EOP_RT'].copy()
                        else:
                            pa_data = usd_xdc_data.copy()
                            eop_data = usd_xdc_data.copy()
                        
                        pa_data = pa_data.dropna(subset=[monthly_value_col])
                        eop_data = eop_data.dropna(subset=[monthly_value_col])
                        
                        # Period average (month_average)
                        if len(pa_data) > 0:
                            pa_data = pa_data.sort_values("month_period")
                            target_month_data = pa_data[
                                pa_data["month_period"].dt.month == self.end_month
                            ]
                            if len(target_month_data) > 0:
                                monthly_averages[currency_code] = float(
                                    target_month_data.iloc[0][monthly_value_col]
                                )
                            ytd_averages[currency_code] = float(pa_data[monthly_value_col].mean())
                        
                        # End of period (closing rate)
                        if len(eop_data) > 0:
                            eop_data = eop_data.sort_values("month_period")
                            target_month_data = eop_data[
                                eop_data["month_period"].dt.month == self.end_month
                            ]
                            if len(target_month_data) > 0:
                                closing_rates[currency_code] = {
                                    "rate": float(target_month_data.iloc[0][monthly_value_col]),
                                    "date": None  # IMF only provides month, not specific dates
                                }
                    
                    # Special handling for EUR: extract from USD data using XDC_EUR indicator
                    # For USA (where USD is the local currency), XDC_EUR tells us EUR per USD
                    usa_data = monthly_df[monthly_df[country_col] == 'USA'].copy()
                    if len(usa_data) > 0:
                        # Get XDC_EUR rates (EUR per 1 USD)
                        xdc_eur_data = usa_data[usa_data[indicator_col] == 'XDC_EUR'].copy()
                        if len(xdc_eur_data) > 0:
                            # Separate period average and end of period
                            if 'TYPE_OF_TRANSFORMATION' in xdc_eur_data.columns:
                                pa_data = xdc_eur_data[xdc_eur_data['TYPE_OF_TRANSFORMATION'] == 'PA_RT'].copy()
                                eop_data = xdc_eur_data[xdc_eur_data['TYPE_OF_TRANSFORMATION'] == 'EOP_RT'].copy()
                            else:
                                pa_data = xdc_eur_data.copy()
                                eop_data = xdc_eur_data.copy()
                            
                            pa_data = pa_data.dropna(subset=[monthly_value_col])
                            eop_data = eop_data.dropna(subset=[monthly_value_col])
                            
                            # Period average for EUR
                            if len(pa_data) > 0:
                                pa_data = pa_data.sort_values("month_period")
                                target_month_data = pa_data[
                                    pa_data["month_period"].dt.month == self.end_month
                                ]
                                if len(target_month_data) > 0:
                                    monthly_averages['EUR'] = float(
                                        target_month_data.iloc[0][monthly_value_col]
                                    )
                                ytd_averages['EUR'] = float(pa_data[monthly_value_col].mean())
                            
                            # End of period for EUR
                            if len(eop_data) > 0:
                                eop_data = eop_data.sort_values("month_period")
                                target_month_data = eop_data[
                                    eop_data["month_period"].dt.month == self.end_month
                                ]
                                if len(target_month_data) > 0:
                                    closing_rates['EUR'] = {
                                        "rate": float(target_month_data.iloc[0][monthly_value_col]),
                                        "date": None
                                    }

            # Combine results
            all_currencies = (
                set(closing_rates.keys())
                | set(monthly_averages.keys())
                | set(ytd_averages.keys())
            )

            results = {}
            for currency in all_currencies:

                # Get closing rate
                closing_rate = cast(
                    float | None, closing_rates.get(currency, {}).get("rate")
                )
                closing_date = cast(
                    str | None, closing_rates.get(currency, {}).get("date")
                )

                # Get averages
                monthly_avg = monthly_averages.get(currency)
                ytd_avg = ytd_averages.get(currency)

                results[currency] = {
                    "Currency": currency,
                    "Closing": closing_rate,
                    "Month_Average": monthly_avg,
                    "YTD_Average": ytd_avg,
                    "Period": closing_date,
                    "has_closing": closing_rate is not None,
                    "has_monthly": monthly_avg is not None,
                    "has_ytd": ytd_avg is not None,
                }

            # Add EUR for ECB source (ECB base currency is EUR, so 1 EUR = 1 EUR)
            if self.sdmx_provider == "ECB" and "EUR" not in results:
                results["EUR"] = {
                    "Currency": "EUR",
                    "Closing": 1.0,
                    "Month_Average": 1.0,
                    "YTD_Average": 1.0,
                    "Period": closing_date if closing_date else None,
                    "has_closing": True,
                    "has_monthly": True,
                    "has_ytd": True,
                }

            # print(f"✓ Successfully processed {len(results)} currencies")

            return results

        except Exception as e:
            print(f"Error retrieving data: {e}")
            import traceback

            traceback.print_exc()
            return {}


class PDF(FPDF):
    def __init__(self, title, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.title = title  # Store the title

    def header(self):
        self.set_font("Helvetica", "B", 16)
        self.cell(0, 10, self.title, new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="C")
        self.ln(10)

    def print_table_header(self, headers, col_width):
        self.set_font("Helvetica", "B", 10)
        self.set_fill_color(230, 230, 230)  # Light grey
        self.set_draw_color(150, 150, 150)  # Middle grey
        for i, header in enumerate(headers):
            align = "L" if i == 0 else "R"
            self.cell(col_width, 10, header, border=1, align=align, fill=True)
        self.cell(0, 0, "", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        self.ln(10)

    def footer(self):
        self.set_y(-15)  # Position 15mm from bottom
        self.set_font("Helvetica", "I", 8)
        date_str = datetime.now().strftime("%Y-%m-%d %H:%M")
        page_text = f"Page {self.page_no()} | Generated on {date_str}"
        self.cell(0, 10, page_text, align="C")

    def add_intro_note(self, text):
        self.set_font("Helvetica", "", 11)
        self.set_text_color(50, 50, 50)  # Optional: darker grey
        self.multi_cell(0, 10, text)
        self.cell(0, 5, "", new_x=XPos.LMARGIN, new_y=YPos.NEXT)  # spacing before table

    def add_footer_note(self, text):
        self.set_font("Helvetica", "", 11)
        self.set_text_color(80, 80, 80)  # Optional: soft grey
        self.cell(0, 5, "", new_x=XPos.LMARGIN, new_y=YPos.NEXT)  # spacing
        self.multi_cell(0, 10, text)

    def table(self, headers, rows):
        self.set_font("Helvetica", "", 10)
        col_width = self.epw / len(headers)  # Equal column width
        self.print_table_header(headers, col_width)

        # Set light grey fill color (RGB: 230, 230, 230)
        # self.set_fill_color(230, 230, 230)

        # for header in headers:
        #    self.cell(col_width, 10, header, border=1, fill=True)
        # self.ln()

        self.set_font("Helvetica", "", 10)

        # Set middle grey border color
        self.set_draw_color(150, 150, 150)

        for row in rows:

            # Check if there's enough space left, else add page and reprint header
            if self.get_y() > self.page_break_trigger - 10:
                self.add_page()
                self.print_table_header(headers, col_width)
                self.set_font("Helvetica", "", 10)

            for i, item in enumerate(row):
                align = "L" if i == 0 else "R"
                # Handle float formatting for non-first columns
                if i != 0:  # and isinstance(item, str):
                    item_clean = str(item).strip().lower()
                    # print(item_clean)
                    if item_clean not in ["", "na", "null", "none"]:
                        try:
                            num = float(item)
                            formatted = f"{num:.8f}"
                            # print(formatted)
                        except ValueError:
                            formatted = item  # fallback if conversion fails
                    else:
                        formatted = ""  # treat as empty
                else:
                    formatted = str(item)

                # Determine borders
                if i == 0:
                    border = "LB"  # Left + Bottom
                elif i == len(row) - 1:
                    border = "RB"  # Right + Bottom
                else:
                    border = "B"  # Only Bottom

                self.cell(col_width, 10, formatted, border=border, align=align)
            self.ln()


# Example usage
if __name__ == "__main__":

    # For the example, enable the use of arguments - Can also be useful for testing purpose ;-)

    # Create the parser
    parser = argparse.ArgumentParser(
        description="ECB - Retrieve Exchange Rates from European Central Bank for a given Period."
    )

    # Add arguments
    parser.add_argument(
        "-y",
        "--year",
        type=int,
        help="Year to get the rates for. Between 2001 and now (optional - default = year of previous month)",
    )
    parser.add_argument(
        "-m",
        "--month",
        type=int,
        help="Month to get the rates for. Between 1 and 12 (optional - default = previous month)",
    )
    parser.add_argument(
        "-s",
        "--start",
        type=int,
        help="First month of the period. Between 1 and 12 (optional - default = 1  == Jan)",
    )
    parser.add_argument(
        "-f",
        "--format",
        type=str,
        help="Output format like csv | xlsx | pdf | txt | json | screen (optional - default = screen)",
    )
    parser.add_argument(
        "-p",
        "--path",
        type=str,
        help="Path for the generated file (optional - default = application folder)",
    )
    parser.add_argument(
        "-s",
        "--source",
        type=str,
        choices=["ECB", "IMF"],
        default="ECB",
        help="Data source: 'ECB' or 'IMF' (default = ECB)",
    )
    parser.add_argument(
        "--daily-key",
        type=str,
        help="Override daily series key for the selected source (optional)",
    )
    parser.add_argument(
        "--monthly-key",
        type=str,
        help="Override monthly series key for the selected source (optional)",
    )
    parser.add_argument(
        "-b",
        "--base-currency",
        type=str,
        help="Convert all rates to be relative to this base currency (optional)",
    )

    # Parse the arguments
    args = parser.parse_args()

    sdmx_year = None if args.year == None else args.year
    sdmx_month = None if args.month == None else args.month
    sdmx_start = None if args.start == None else args.start
    source = args.source.upper()
    override_daily_key = args.daily_key
    override_monthly_key = args.monthly_key
    base_currency = args.base_currency.upper() if args.base_currency else None

    # Define provider/dataset/key mapping for supported sources
    SOURCE_MAP = {
        "ECB": {
            "provider": "ECB",
            "dataset": "EXR",
            "daily_key": "D..EUR.SP00.A",
            "monthly_key": "M..EUR.SP00.A",
        },
        "IMF": {
            # IMF SDMX registry uses 'IMF_DATA' for the client in the sdmx library
            # Use the IMF Exchange Rates dataflow 'ER' by default (contains period average and end-of-period rates)
            "provider": "IMF_DATA",
            "dataset": "ER",
            "daily_key": None,
            "monthly_key": None,
        },
    }

    if source not in SOURCE_MAP:
        print(f"Unknown source '{source}', defaulting to ECB")
        source = "ECB"

    src_info = SOURCE_MAP[source]
    provider = src_info.get("provider")
    dataset = src_info.get("dataset")
    daily_key = override_daily_key if override_daily_key is not None else src_info.get("daily_key")
    monthly_key = override_monthly_key if override_monthly_key is not None else src_info.get("monthly_key")
    output: str | None = (
        "screen" if args.format == None else args.format.strip().lower()
    )

    #print(f"{datetime.now()}: Retrieving exchange rates from the selected source for the provided arguments...")
    ecbexr = Ecbexr(sdmx_year, sdmx_month, sdmx_start, provider=provider, dataset=dataset, daily_key=daily_key, monthly_key=monthly_key)

    # If IMF is selected and no monthly key is provided, we'll fetch the dataset without a series key
    # which retrieves all series for the dataset (may be large). This mirrors ECB behaviour
    # where the default keys include all currencies.
    if source == "IMF" and monthly_key is None:
        print(
            "IMF source selected with no monthly key: the script will fetch all series in the IMF dataset (may be large)."
        )
    rates_dict = ecbexr.rates()

    # If no data was retrieved, exit gracefully instead of failing later
    if not rates_dict:
        print("No rates available. Exiting.")
        sys.exit(0)
    
    # Apply base currency conversion if specified
    if base_currency:
        if base_currency in rates_dict:
            # Create a copy of the base_rate_dict to avoid modifying it as we iterate
            import copy
            base_rate_dict = copy.copy(rates_dict[base_currency])
            numeric_columns = ['Closing', 'Month_Average', 'YTD_Average']
            
            for currency_code in rates_dict:
                for col in numeric_columns:
                    if col in rates_dict[currency_code]:
                        current_val = rates_dict[currency_code][col]
                        base_val = base_rate_dict.get(col)
                        
                        # If both values exist and base_val is not zero, divide
                        if current_val is not None and base_val is not None and base_val != 0:
                            rates_dict[currency_code][col] = current_val / base_val
                        elif base_val == 0 or base_val is None:
                            # If base rate is zero or None, set to zero
                            rates_dict[currency_code][col] = 0
        else:
            print(f"Warning: Base currency '{base_currency}' not found in retrieved data. No conversion applied.")
            print(f"Available currencies: {', '.join(sorted(rates_dict.keys()))}")

    if args.path is not None:
        base_dir = Path(args.path)
    else:
        # __file__ = chemin du script, .parent = répertoire du script
        base_dir = Path(__file__).parent
    
    # Build filename with source and base currency information
    base_currency_str = base_currency if base_currency else "original"
    filename = base_dir / f"ExchangeRates_{source}_{base_currency_str}_{ecbexr.end_year}-{ecbexr.end_month:02d}"

    # Convert dictionary to table
    #print(f"{datetime.now()}: Preparing data...")

    # Columns to exclude
    exclude_columns = ["has_closing", "has_monthly", "has_ytd", "month_end_date"]

    # Prepare headers by filtering keys from the first row
    headers = [
        key for key in next(iter(rates_dict.values())) if key not in exclude_columns
    ]

    # Prepare rows
    rows = [
        [rates_dict[outer_key][key] for key in headers]
        for outer_key in sorted(rates_dict.keys())
    ]

    #print(f"{datetime.now()}: Generating output...")
    if output == "pdf":

        title = (
            f"Exchange rates for the period: {ecbexr.end_year}-{ecbexr.end_month:02d}"
        )
        pdf = PDF(title)
        pdf.add_page()

        # Add intro text before the table
        pdf.add_intro_note("Currency rates from the European Central Bank.")
        pdf.table(headers, rows)
        pdf.add_footer_note(
            "Note: All values are rounded to 8 decimal places. Missing entries are left blank."
        )

        # Save the PDF
        pdf.output(f"{filename}.pdf")

        print(
            f"{filename}.pdf generated ..."
        )

    elif output == "csv":

        # Write to CSV
        with open(
            f"{filename}.csv",
            "w",
            newline="",
            encoding="utf-8",
        ) as f:
            writer = csv.writer(f)
            writer.writerow(headers)  # write header
            writer.writerows(rows)  # write data rows

        print(
            f"{filename}.csv generated ..."
        )

    elif output == "txt":
        # Open (or create) a file in write mode
        with open(
            f"{filename}.txt",
            "w",
            encoding="utf-8",
        ) as file:
            file.write(
                f"\n Rates from ECB for period {ecbexr.end_year}-{ecbexr.end_month:02d} using {ecbexr.start_month:02d} as starting month"
            )
            file.write("\n")
            file.write("=" * 70)
            file.write("\n")
            file.write("\n")
            file.write(tabulate(rows, headers=headers, tablefmt="grid"))

        print(
            f"{filename}.txt generated ..."
        )

    elif output == "json":
        # Open (or create) a file in write mode
        with open(
            f"{filename}.json",
            "w",
            encoding="utf-8",
        ) as file:
            pprint.pprint(rates_dict, stream=file)

        print(
            f"{filename}.json generated ..."
        )

    elif output == "xlsx":

        # Create workbook and sheet
        wb = Workbook()
        ws = wb.active
        if ws is not None:
            ws.title = f"Exr_{ecbexr.end_year}-{ecbexr.end_month:02d}"

            # Write headers
            ws.append(headers)

            # Write rows
            for outer_key in sorted(rates_dict.keys()):
                row = [rates_dict[outer_key][key] for key in headers]
                ws.append(row)

            # Save file
            wb.save(f"{filename}.xlsx")

            print(
                f"{filename}.xlsx generated ..."
            )

    else:

        print(
            f"\n Retrieving rates from {source} for period {ecbexr.end_year}-{ecbexr.end_month:02d} using {ecbexr.start_month:02d} as starting month",
            "\n",
            "=" * 70,
        )
        print("\n\n Rates:\n", "-" * 6)

        # Print table
        print(tabulate(rows, headers=headers, tablefmt="grid"))
